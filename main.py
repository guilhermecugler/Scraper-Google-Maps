import time

import customtkinter
import geopandas as gpd
import json
import numpy as np
import os
import pandas as pd
import pyautogui
import pywhatkit as py
import re
import requests
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import (
    expected_conditions as EC,
    expected_conditions,
)
from selenium.webdriver.support.ui import WebDriverWait
from tkintermapview import TkinterMapView
import threading
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.alignment import Alignment
from openpyxl.styles import PatternFill
from threading import Thread
from threading import Event

customtkinter.set_default_color_theme("blue")

class StatusBar(customtkinter.CTkFrame):
    def __init__(self, master):
        customtkinter.CTkFrame.__init__(self, master)
         
        self.label = customtkinter.CTkLabel(self, text = "")
        self.label.pack(side = customtkinter.LEFT)
        self.pack(fill=customtkinter.X, side = customtkinter.BOTTOM)
     
    def setText(self, newText):
        self.label.configure(text = newText)
 
    def clearText(self):
        self.label.configure(text = "")

class App(customtkinter.CTk):

    APP_NAME = "Captação de Cliente Google Maps"
    WIDTH = 800
    HEIGHT = 500

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)


        self.event = Event()
        self.title(App.APP_NAME)
        self.geometry(str(App.WIDTH) + "x" + str(App.HEIGHT))
        self.minsize(App.WIDTH, App.HEIGHT)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.bind("<Command-q>", self.on_closing)
        self.bind("<Command-w>", self.on_closing)
        self.createcommand('tk::mac::Quit', self.on_closing)

        self.marker_list = []
        self.current_position = ()

        # ============ create two CTkFrames ============

        self.grid_columnconfigure(0, weight=0)
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)

        self.frame_left = customtkinter.CTkFrame(master=self, width=150, corner_radius=0, fg_color=None)
        self.frame_left.grid(row=0, column=0, padx=0, pady=0, sticky="nsew")

        self.frame_right = customtkinter.CTkFrame(master=self, corner_radius=0)
        self.frame_right.grid(row=0, column=1, rowspan=1, pady=0, padx=0, sticky="nsew")
        
        self.CTkFrame = customtkinter.CTkFrame(master=self, width=150, corner_radius=0)
        self.status_bar = StatusBar(self.CTkFrame)
        self.CTkFrame.grid(padx=40, column=1, pady=10)
        # ============ frame_left ============

        self.frame_left.grid_rowconfigure(2, weight=1)

        self.button_1 = customtkinter.CTkButton(master=self.frame_left,
                                                text="Marcar Local",
                                                command=self.set_marker_event)
        self.button_1.grid(pady=(20, 0), padx=(20, 20), row=0, column=0)

        self.button_2 = customtkinter.CTkButton(master=self.frame_left,
                                                text="Limpar",
                                                command=self.clear_marker_event)
        self.button_2.grid(pady=(20, 0), padx=(20, 20), row=1, column=0)

        self.entry_keywords = customtkinter.CTkEntry(master=self.frame_left,
                                            placeholder_text="Palavras-chave")
        self.entry_keywords.grid(row=3, column=0, sticky="n", padx=(20, 20), pady=5)

        self.entry_radius = customtkinter.CTkEntry(master=self.frame_left,
                                            placeholder_text="Distância em metros")
        self.entry_radius.grid(row=3, column=0, sticky="we", padx=(20, 20), pady=(40,  0))

        self.button_search_clients = customtkinter.CTkButton(master=self.frame_left,
                                                text="Buscar Clientes",
                                                command=self.threadBuscar)
        self.button_search_clients.grid(pady=(0, 0), padx=(20, 20), row=4, column=0)

        self.button_cancelar = customtkinter.CTkButton(master=self.frame_left,
                                                text="Cancelar",
                                                command=self.fecharThread)
        self.button_cancelar.grid(pady=(60, 0), padx=(20, 20), row=4, column=0)

        self.button_message_instagram = customtkinter.CTkButton(master=self.frame_left,
                                                text="Mensagem Instagram",
                                                command=self.threadInsta)
        self.button_message_instagram.grid(pady=(0, 0), padx=(20, 20), row=5, column=0)

        self.button_message_whatsapp = customtkinter.CTkButton(master=self.frame_left,
                                                text="Mensagem WhatsApp",
                                                command=self.threadZap)
        self.button_message_whatsapp.grid(pady=(80, 0), padx=(20, 20), row=5, column=0)



        self.map_label = customtkinter.CTkLabel(self.frame_left, text="Tipo de mapa:", anchor="w")
        self.map_label.grid(row=6, column=0, padx=(20, 20), pady=(20, 0))
        self.map_option_menu = customtkinter.CTkOptionMenu(self.frame_left, values=["OpenStreetMap", "Google normal", "Google satellite"],
                                                                       command=self.change_map)
        self.map_option_menu.grid(row=7, column=0, padx=(20, 20), pady=(10, 0))

        self.appearance_mode_label = customtkinter.CTkLabel(self.frame_left, text="Tema:", anchor="w")
        self.appearance_mode_label.grid(row=8, column=0, padx=(20, 20), pady=(20, 0))
        self.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(self.frame_left, values=["Light", "Dark", "System"],
                                                                       command=self.change_appearance_mode)
        self.appearance_mode_optionemenu.grid(row=9, column=0, padx=(20, 20), pady=(10, 20))

        # ============ frame_right ============

        self.frame_right.grid_rowconfigure(1, weight=1)
        self.frame_right.grid_rowconfigure(0, weight=0)
        self.frame_right.grid_columnconfigure(0, weight=1)
        self.frame_right.grid_columnconfigure(1, weight=0)
        self.frame_right.grid_columnconfigure(2, weight=1)

        self.map_widget = TkinterMapView(self.frame_right, corner_radius=0)
        self.map_widget.grid(row=1, rowspan=1, column=0, columnspan=3, sticky="nswe", padx=(0, 0), pady=(0, 0))
        self.map_widget.add_left_click_map_command(self.left_click_event)

        self.entry = customtkinter.CTkEntry(master=self.frame_right,
                                            placeholder_text="Digite um local")
        self.entry.grid(row=0, column=0, sticky="we", padx=(12, 0), pady=12)
        self.entry.bind("<Return>", self.search_event)

        self.button_5 = customtkinter.CTkButton(master=self.frame_right,
                                                text="Buscar",
                                                width=90,
                                                command=self.search_event)
        self.button_5.grid(row=0, column=1, sticky="w", padx=(12, 0), pady=12)

        # Set default values
        self.map_widget.set_address("Rio de Janeiro")
        self.map_option_menu.set("OpenStreetMap")
        self.appearance_mode_optionemenu.set("Dark")

    def left_click_event(map_widget):
         current_position = map_widget.get_position()

    def handleStatusBar(self, texto):
        self.status_bar.setText(texto)

    def search_event(self, event=None):
        self.map_widget.set_address(self.entry.get())

    def set_marker_event(self):
        self.current_position = self.map_widget.get_position()
        self.marker_list.append(self.map_widget.set_marker(self.current_position[0], self.current_position[1]))

    def clear_marker_event(self):
        for marker in self.marker_list:
            marker.delete()

    def change_appearance_mode(self, new_appearance_mode: str):
        customtkinter.set_appearance_mode(new_appearance_mode)

    def change_map(self, new_map: str):
        if new_map == "OpenStreetMap":
            self.map_widget.set_tile_server("https://a.tile.openstreetmap.org/{z}/{x}/{y}.png")
        elif new_map == "Google normal":
            self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=m&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)
        elif new_map == "Google satellite":
            self.map_widget.set_tile_server("https://mt0.google.com/vt/lyrs=s&hl=en&x={x}&y={y}&z={z}&s=Ga", max_zoom=22)

    def on_closing(self, event=0):
        self.destroy()

    def start(self):
        self.mainloop()

    def find_locations(self, search_url, api_key):
                
                #lista de listas para todos os dados
                final_data = []

                #while loop para solicitar e analisar os arquivos JSON solicitados
                while True:
                    if self.event.is_set():
                        self.event.clear()
                        self.handleStatusBar(texto=f"Cancelado com sucesso!")
                        
                        break
                    # print(search_url)
                    time.sleep(2)
                    respon = requests.get(search_url)
                    jj = json.loads(respon.text)
                    results = jj['results']
                    #analise todas as informa��es necess�rias
                    for result in results:
                        if self.event.is_set():
                            self.event.clear()
                            self.handleStatusBar(texto=f"Cancelado com sucesso!")
                            
                            break
                        name = result['name']
                        place_id = result ['place_id']
                        lat = result['geometry']['location']['lat']
                        longi = result['geometry']['location']['lng']
                        try:
                            rating = result['rating']
                        except:
                            pass
                        types = result['types']
                        data = [name, place_id, lat, longi, rating, types]
                        final_data.append(data)
                        # print(data)
                    
                    #se houver uma pr�xima p�gina, o loop ser� reiniciado com uma url atualizada
                    #se n�o houver pr�xima p�gina, o programa grava em um csv e salva em df    
                    if 'next_page_token' not in jj:
                        labels = ['Place Name','ID_Field', 'Latitude', 'Longitude', 'Rating', 'Tags']
                        location_df = pd.DataFrame.from_records(final_data, columns=labels)
                        #location_df.to_csv('location.csv')
                        break
                    else:
                        next_page_token = jj['next_page_token']
                        # print('looooooop')
                        search_url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?key='+str(api_key)+'&pagetoken='+str(next_page_token)
                        
                return(final_data, location_df)
    
    def find_details(self, final_data, api_key):
                
                final_detailed_data =[]
                #Usa o ID exclusivo de cada local para usar outra solicita��o de API para obter informa��es de telefone e site de cada empresa.
                for places in final_data:
                    if self.event.is_set():
                        self.event.clear()
                        self.handleStatusBar(texto=f"Cancelado com sucesso!")
                        
                        break
                    id_field = places[1]
                    self.handleStatusBar(texto=f"Encontrados {len(final_data)} prospects, puxando dados: \n{places[0]}")
                    self.update_idletasks()
                    req_url = 'https://maps.googleapis.com/maps/api/place/details/json?place_id='+str(id_field)+'&fields=name,formatted_phone_number,website&key='+str(api_key)
                    respon = requests.get(req_url)
                    jj = json.loads(respon.text)
                    results = jj['result']

                    identification = id_field
                    try:
                        phone = results["formatted_phone_number"]
                    except KeyError:
                        continue
                    try:
                        website = results["website"]
                    except KeyError:
                        continue
                    title = results["name"]
                    detailed_data = [title, identification, phone, website]
                    final_detailed_data.append(detailed_data)




                columns = ["Business", "ID_Field","Phone", "Website"]
                details_df = pd.DataFrame.from_records(final_detailed_data, columns=columns)
                details_df.to_csv('further_details.csv', index=False)
                
                return details_df
    
    def join_data(self, details_df,location_df,keyword):

                final_sheet = location_df.join(details_df.set_index('ID_Field'), on='ID_Field')

                final_sheet.to_csv(str(keyword) + ".csv", index=False)

                # print(final_sheet)

                return final_sheet
    
    def csv_to_point(self, non_spatial_data):
                #crie o geodataframe e exporte-o como um arquivo de ponto
                del non_spatial_data['Tags']
                spatial_df = gpd.GeoDataFrame(non_spatial_data, geometry=gpd.points_from_xy(non_spatial_data.Longitude, non_spatial_data.Latitude))
                #spatial_df.to_csv("point_data.csv")
                # print(spatial_df)
                #spatial_df.to_file("point_data.shp")

                #create a projection file that corresponds to where data was taken from
                #prj = open("point_data.prj", "w")
                epsg = 'GEOGCS["GCS_WGS_1984",DATUM["D_WGS_1984",SPHEROID["WGS_1984",6378137,298.257223563]],PRIMEM["Greenwich",0],UNIT["Degree",0.017453292519943295]]'
                #prj.write(epsg)
                #prj.close()

                return(spatial_df)
    
    def planilha(self):
        path = r"./"#COLOCAR O LOCAL ONDE EST� OS EXCELS SALVOS
        # use compreens�o de lista para criar uma lista de arquivos csv
        csv_files = [file for file in os.listdir(path) if file.endswith('.csv')]
        csv_files.remove('further_details.csv')
        # cria uma lista vazia para armazenar os DataFrames
        df_list = []

        # iterar pelos arquivos csv
        for csv_file in csv_files:
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                
                break
            # l� o arquivo csv atual em um DataFrame
            df = pd.read_csv(os.path.join(path, csv_file))
            # adiciona uma nova coluna 'file_name' com o nome do arquivo csv atual
            df['name'] = os.path.basename(csv_file)
            # anexa o DataFrame atual � lista
            df_list.append(df)

        # concatenar todos os DataFrames em um �nico DataFrame

        merged_df = pd.concat(df_list)

        # grava o DataFrame mesclado em um novo arquivo csv
        merged_df = merged_df.drop(columns=['ID_Field', 'Latitude', 'Longitude', 'Tags', 'name'], axis=1)
        
        merged_df = merged_df.rename(columns={"Place Name": "Nome", "Rating": "Avaliação", "Phone": "Telefone"})

        try:
            planilha_anterior = pd.read_excel("all.xlsx")

            frames = [merged_df, planilha_anterior]
    
            merged_df = pd.concat(frames)
        except:
            pass
        

        merged_df = merged_df.drop_duplicates()
        merged_df.to_excel("all.xlsx", index=False)



    insta_message = []
    face_message = []
    others_websites = []
    find_instagram = []
    usuarios_instagram = []
    notfind_instagram = []
    list_instagram = []
    

    def threadBuscar(self):
        thread = Thread(target=self.search_clients, args=(self.event,))
        thread.daemon = True
        thread.start()

    def threadZap(self):
        thread = Thread(target=self.send_whatsapp_message, args=(self.event,))
        thread.start()

    def threadInsta(self):
        thread = Thread(target=self.send_instagram_message, args=(self.event,))
        thread.start()

    def fecharThread(self):
        self.event.set()
        
        print("Clicou fechar ")
        print(self.event.is_set())



    def search_clients(self, event):
        self.event.clear()
        print(self.event.is_set())
        

        # report a message
        self.handleStatusBar(texto=f"Iniciando buscas, aguarde...")

        print('Iniciando busca assincrona...')

        api_key = 'AIzaSyAQBPAC422_rwNTDvfM8tq2Opma5VCa12A'
        search_radius = self.entry_radius.get()
        
        coords = ','.join([str(value) for value in self.current_position])
        
        keywords = self.entry_keywords.get().split(",")

        palavras_chave = keywords

        for keyword in palavras_chave:
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                break
            request_url = 'https://maps.googleapis.com/maps/api/place/nearbysearch/json?location='+coords+'&radius='+str(search_radius)+'&keyword='+str(keyword)+'&key='+str(api_key)
            
            #encontre os locais dos estabelecimentos desejados no google maps
            final_data, location_df = self.find_locations(request_url, api_key)
            
            #encontre site, telefone e avalia��es de estabelecimentos
            details_df = self.find_details(final_data, api_key)

            #junte os dois dataframes para ter um produto final
            non_spatial_data = self.join_data(details_df,location_df,keyword)
            

            #c
            spatial_df = self.csv_to_point(non_spatial_data)

        

        self.planilha()
        self.handleStatusBar(texto="Salvo na planilha!")
        self.update_idletasks()
        self.organize_sheets()
        self.organizar_planilha()

        self.insta_message.clear()
        self.face_message.clear()
        self.others_websites.clear()
        self.find_instagram.clear()
        self.usuarios_instagram.clear()
        self.notfind_instagram.clear()
        self.list_instagram.clear()
        del final_data, location_df, details_df, spatial_df
        csv_files = [file for file in os.listdir('./') if file.endswith('.csv')]
        
        for csv_file in csv_files:
            os.remove(csv_file)




    def organize_sheets(self):
        planilha = pd.read_excel("all.xlsx")
        planilha = planilha[~planilha['Nome'].str.contains('Magazine|Kalunga|renner|Americanas|Americana|americanas|americana|Mercado Livre|mercado livre|shopee|Shopee|Renner', na=False)]
        planilha = planilha.drop_duplicates(subset='Nome')
        
        #Instagram
        for i in range(len(planilha)):
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                
                return
            self.handleStatusBar(texto=f"Procurando instagram dos prospects {i}")

            first_row = planilha.iloc[i]
            url = re.search(r'(www\.instagram\.com)', str(first_row["Website"]))

            if url:
                self.insta_message.append(first_row["Website"])
                planilha["Website"].replace(to_replace=first_row["Website"],value="Done",inplace=True)
            else:
                pass
        #Facebook
        for i in range(len(planilha)):
            self.handleStatusBar(texto=f"Procurando facebook dos prospects {i}")

            first_row = planilha.iloc[i]
            url = re.search(r'(www\.facebook\.com)', str(first_row["Website"]))

            if url:
                self.face_message.append(first_row["Website"])
                planilha["Website"].replace(to_replace=first_row["Website"],value="Done",inplace=True)

            else:
                pass

        #Other Websites
        for i in range(len(planilha)):
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return

            first_row = planilha.iloc[i]
            url2 = re.search(r'.www.', str(first_row["Website"]))
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return
            if url2:
                if self.event.is_set():
                    self.event.clear()
                    self.handleStatusBar(texto=f"Cancelado com sucesso!")

                    return
                self.others_websites.append(first_row["Website"])
                planilha["Website"].replace(to_replace=first_row["Website"],value="Done",inplace=True)
            else:
                if self.event.is_set():
                    self.event.clear()
                    self.handleStatusBar(texto=f"Cancelado com sucesso!")

                    return
                pass
        print("Procurando por Instagram dos clientes")
        for g in range(len(self.others_websites)):
            self.handleStatusBar(texto=f"Verificando sites/contatos {g}/{len(self.others_websites)}")

            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return
            try:
                response = requests.get(g)

                if response.status_code == 200:
                    if self.event.is_set():
                        self.event.clear()
                        self.handleStatusBar(texto=f"Cancelado com sucesso!")

                        return
                    website = response.text
                    
                    url = re.search(r'www\.instagram\.com\/[A-Za-z0-9]+\/', website)

                    if url:
                        if self.event.is_set():
                            self.event.clear()
                            self.handleStatusBar(texto=f"Cancelado com sucesso!")

                            return
                        self.find_instagram.append(url)
                        matched_string2 = url.group()
                        #planilha["Website"].replace(to_replace=first_row["Website"],value="Done",inplace=True)

                    else:
                        if self.event.is_set():
                            self.event.clear()
                            self.handleStatusBar(texto=f"Cancelado com sucesso!")

                            return
                        self.notfind_instagram.append(g)
                        
                        pass    
                else:
                    print("Site não acessivel", g)
                    continue
            except:
                continue

        for h in self.find_instagram:
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                
                return
            self.handleStatusBar(texto=f"Verificando instagram válido do prospect {h}")
            url = re.search(r'www\.instagram\.com\/[A-Za-z0-9]+\/', str(h))

            if url:

                matched_string = url.group()
                self.insta_message.append(matched_string)
            else:
                print("error")
        

        for i in self.insta_message:
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                return
            instagram = i.replace('https://www.instagram.com/', '')
            instagram = instagram.replace('http://www.instagram.com/', '')
            instagram = instagram.replace('www.instagram.com/', '')
            instagram = instagram.replace('/', '')
            instagram = instagram.replace('?hl=pt-br', '')
            self.handleStatusBar(texto=f"Instagram válido: {instagram}")

            self.list_instagram.append(instagram)
        
        # jsonInstagram = json.dumps(self.list_instagram)
        # jsonFile = open("lista_instagram.json", "w")
        # jsonFile.write(jsonInstagram)
        # jsonFile.close()
        self.handleStatusBar(texto=f"Lista de Instagram salva!")


        nada_encontrado = []

        planilha = pd.read_excel("all.xlsx")
        planilha = planilha[~planilha['Nome'].str.contains('Magazine|Kalunga|renner|Americanas|Americana|americanas|americana|Mercado Livre|mercado livre|shopee|Shopee|Renner', na=False)]
        planilha = planilha.drop_duplicates(subset='Nome')
        nada_encontrado = []

        for i in range(len(planilha)):
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return
            first_row = planilha.iloc[i]
            if first_row["Telefone"] != "Done":
                if first_row["Website"] != "Done":
                    nada_encontrado.append(first_row["Nome"])
                    #colocar em um dicionario com as principais informa��es -> fazer alguma coisa para achar esses locais

        list_numbers = []
        for i in range(len(planilha)):
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")

                return
            first_row = planilha.iloc[i]
            if not pd.isnull(first_row["Telefone"]):
                list_numbers.append(first_row["Telefone"])
                planilha["Telefone"].replace(to_replace=first_row["Telefone"],value="Done",inplace=True)

        list_numbers = [re.sub(r'[^\d]+', '', string) for string in list_numbers]
        list_numbers = [item for item in list_numbers if item]

        # jsonNumbers = json.dumps(list_numbers)
        # jsonFileNumbers = open("lista_whatsapp.json", "w")
        # jsonFileNumbers.write(jsonNumbers)
        # jsonFileNumbers.close()


        if self.event.is_set():
            self.handleStatusBar(texto=f"Cancelado com sucesso!")
        else:
            self.handleStatusBar(texto=f"Dados salvos!")


    def organizar_planilha(nome_planilha):
        # isFile = os.path.isfile(nome_planilha)
        # if(isFile == False):
        #     writer = pd.ExcelWriter(nome_planilha, engine='xlsxwriter')
        #     writer.close()
        nome_planilha = "all.xlsx"
        writer = pd.ExcelWriter(nome_planilha, engine='openpyxl', if_sheet_exists='replace', mode='a')

        workbook  = writer.book



        sheet = workbook['Sheet1']

        for column_cells in sheet.columns:
            new_column_length = max(len(str(cell.value)) for cell in column_cells)
            new_column_letter = (get_column_letter(column_cells[0].column))
            if new_column_length > 0:
                sheet.column_dimensions[new_column_letter].width = new_column_length*1.23
            
        for row in range(1,sheet.max_row+1):
            for col in range(1,sheet.max_column+1):
                cell=sheet.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row%2==0:
                    start_color='BFBFBF' 
                else :
                    start_color='00FFFFFF'
                cell.fill = PatternFill(start_color=start_color, fill_type="solid")

            
        sheet.column_dimensions['A'].width = 60
        sheet.column_dimensions['C'].width = 40
        sheet.column_dimensions['E'].width = 60


        writer.close()

        return "Planilha Organizada"

    def message_instagram(self, username, password, lista_instagram, message):
        if self.event.is_set():
            self.event.clear()
            self.handleStatusBar(texto=f"Cancelado com sucesso!")
            return
        options = webdriver.ChromeOptions()
        # options.add_argument("--headless=new")

        driver = webdriver.Chrome(options=options)

        self.username = username
        self.password = password
        self.user = lista_instagram
        self.message = message
        self.base_url = 'https://www.instagram.com/'
        self.direct_url = 'https://www.instagram.com/direct/new/?hl=pt-br'
        self.bot = driver
        self.login()

    def login(self):
        self.bot.get(self.base_url)

        enter_username = WebDriverWait(self.bot, 20).until(
            expected_conditions.presence_of_element_located((By.NAME, 'username')))
        enter_username.send_keys(self.username)
        enter_password = WebDriverWait(self.bot, 20).until(
            expected_conditions.presence_of_element_located((By.NAME, 'password')))
        enter_password.send_keys(self.password)
        enter_password.send_keys(Keys.RETURN)
        time.sleep(5)

        # # first pop-up
        # self.bot.find_element(By.XPATH,
        #                     "(//div[normalize-space()='Agora não'])[1]").click()
        # time.sleep(5)



        # direct
        self.bot.get(self.direct_url)
        time.sleep(3)
        if self.event.is_set():
            self.event.clear()
            self.handleStatusBar(texto=f"Cancelado com sucesso!")
            return
        # 2nd pop-up
        self.bot.find_element(By.XPATH,
                            "(//button[normalize-space()='Agora não'])[1]").click()
        time.sleep(5)
        for i in self.user:
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                break
            self.handleStatusBar(f"Enviando mensagem para instagram: {i}")

            
            # enter the username
            self.bot.find_element(By.XPATH,
                                "//input[@placeholder='Pesquisar...']").send_keys(i)
            time.sleep(3)

            # click on the username
            self.bot.find_element(By.XPATH,
                                f"//span[text()='{i}']").click()
            time.sleep(4)

            # next button
            self.bot.find_element(By.XPATH,
                                "(//div[normalize-space()='Avançar'])[1]").click()
            time.sleep(4)

            # click on message area
            send = self.bot.find_element(By.XPATH,
                                        "//textarea[@placeholder='Mensagem...']")

            # types message
            send.send_keys(self.message)
            time.sleep(1)

            # send message
            send.send_keys(Keys.RETURN)
            time.sleep(2)

            # clicks on direct option or pencil icon
            self.bot.find_element(By.XPATH,
                                "//*[name() = 'svg'][contains(@aria-label, 'Nova mensagem')]").click()
            time.sleep(4)
    
    def send_instagram_message(self, event):
        if self.event.is_set():
            self.event.clear()
            self.handleStatusBar(texto=f"Cancelado com sucesso!")
            return
        f = open('lista_instagram.json')   
        g = open('credenciais_instagram.json')      
        h = open('mensagem.json', encoding='utf-8')      

        credencial_instagram = json.load(g)
        usuario = credencial_instagram['usuario']
        senha = credencial_instagram['senha']

        lista_instagram = json.load(f)

        message = f"{json.load(h)}"
        self.message_instagram(usuario, senha, lista_instagram, message)
    
        # when our program ends it will show "done".
        self.handleStatusBar("Mensagens enviadas pelo Instagram")    
    # def send_instagram_message(self, user, password, message_):
    #     self.message_instagram(user, password, self.list_instagram, message_)
    
    #     # when our program ends it will show "done".
    #     self.handleStatusBar("Mensagens enviadas pelo Instagram")

    def send_whatsapp_message(self, event):
        f = open('lista_whatsapp.json')   
        h = open('mensagem.json', encoding='utf-8')      

        list_numbers = json.load(f)
        message = f"{json.load(h)}"

        print("Iniciando envio de Mensagens pelo WhatsApp")
        for i in range(len(list_numbers)):
            if self.event.is_set():
                self.event.clear()
                self.handleStatusBar(texto=f"Cancelado com sucesso!")
                return
            py.sendwhatmsg_instantly(f'+55{list_numbers[i]}', message, tab_close=True)
            self.handleStatusBar(f'Enviando mensagem para WhatsApp {list_numbers[i]}')

        
        self.handleStatusBar("Mensagens enviadas pelo WhatsApp.")

if __name__ == "__main__":
    app = App()
    app.start()
    app.mainloop()